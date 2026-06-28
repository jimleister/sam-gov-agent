#!/usr/bin/env python3
"""
SAM.gov Daily Scan (Get Opportunities Public API v2)

Goals (updated for Weston Trolley + WEXMAC logistics focus)
- Pull opportunities posted in the last 72 hours, Active only, for notice types:
  Sources Sought, Presolicitation, Solicitation, Combined Synopsis/Solicitation
- Use OR-logic across structured “signal families” (state, NAICS, PSC):
  run multiple API searches -> union -> dedupe by noticeId
- Match keywords LOCALLY (title + description). No API q= keyword searches.
- Rank Top opportunities by a FEASIBILITY ratio derived from:
    Profitability (higher is better) vs (Complexity + Overhead) (lower is better)
  Relevance (number of matched signals) is a light tiebreaker.
- Top 5–10 should be drawn from *all* matches (keywords/NAICS/PSC/state/org/etc),
  with no “core keyword” gating/boosting.
- Shortlist 10–20: next-best items (including state-only matches).

Email output
- Writes ./email_draft.txt and prints to stdout.
- Includes To and Cc headers (configurable).
- Does NOT include any API keys.
- Optional: can send via Gmail SMTP if you set SEND_EMAIL=1 and SMTP env vars (see below).

Secrets
- Requires SAM_API_KEY in environment.

Optional SMTP sending (Gmail)
- Set these env vars (recommended in ~/.zprofile), and set SEND_EMAIL=1:
  SMTP_HOST="smtp.gmail.com"
  SMTP_PORT="587"
  SMTP_USER="westontrolley@gmail.com"
  SMTP_PASS="<Gmail App Password>"   # NOT your normal password
  FROM_EMAIL="westontrolley@gmail.com"
"""

from __future__ import annotations

import os
import sys
import re
import json
import time
import csv
import html
import mimetypes
from pathlib import Path
import datetime as dt
from dataclasses import dataclass, field
from typing import Any, Dict, List, Tuple, Optional

import requests

# Optional SMTP (only used if SEND_EMAIL=1)
import smtplib
from email.message import EmailMessage


# -----------------------------
# CONFIG
# -----------------------------

SAM_SEARCH_URL = "https://api.sam.gov/prod/opportunities/v2/search"

NOTICE_TYPES = ["r", "p", "o", "k"]  # Sources Sought, Presolicitation, Solicitation, Combined Synopsis/Solicitation
POSTED_WINDOW_HOURS = 72
ACTIVE_ONLY = True

# LOCAL keyword filtering only (no API q=)
# Weston Trolley target universe: passenger transportation/trolley/shuttle plus
# WEXMAC logistics support: cargo movement, vehicle rental/driver services, warehouse,
# loading/unloading, customs/clearance, water ferry/taxi, base operations/life support,
# construction/material-handling equipment, lodging/catering, medical logistics, force
# protection, communications support, food/rations/water. VOSB/small-business language
# is handled below as set-aside signals and scoring boosters.
KEYWORDS = [
    # core trolley / shuttle / bus operations
    "trolley", "trolley service", "historic trolley", "streetcar",
    "shuttle", "shuttle service", "bus shuttle", "courtesy shuttle",
    "circulator", "downtown circulator", "visitor shuttle", "park shuttle",
    "transit operations", "transportation services", "passenger transportation",
    "ground transportation", "surface transportation", "fixed route",
    "bus service", "motor coach", "motorcoach", "charter bus", "coach bus",

    # municipal / campus / event / tourism use cases
    "event transportation", "special event transportation", "festival shuttle",
    "campus shuttle", "employee shuttle", "commuter shuttle",
    "tour transportation", "sightseeing", "visitor transportation",
    "concessionaire", "concession", "recreation area", "national park",
    "theme park", "fairgrounds", "parking shuttle", "lot shuttle",

    # accessibility / community transport
    "paratransit", "microtransit", "demand response", "non-emergency transportation",
    "mobility services", "senior transportation", "ada transportation",

    # fleet, drivers, dispatch, maintenance
    "vehicle operator", "driver services", "bus operator", "dispatch",
    "fleet maintenance", "vehicle maintenance", "preventive maintenance",
    "bus maintenance", "transit vehicle", "vehicle leasing", "vehicle rental",

    # WEXMAC logistics and transportation services
    "cargo truck", "cargo van", "covered truck", "flatbed", "stake truck",
    "semi truck", "tractor trailer", "reefer van", "refrigerated van",
    "vehicle rental", "vehicle leasing", "rental with driver", "rental without driver",
    "sedan with driver", "van with driver", "van without driver",
    "airport transfer", "airport transfers", "personnel logistic movement",
    "personnel logistics movement", "logistic movement support", "PLMS",
    "warehouse", "warehousing", "general warehouse", "hazmat warehouse",
    "portable warehouse", "storage services", "loading", "unloading",
    "loading/unloading", "customs clearance", "customs duty", "bill of lading",
    "freight", "drayage", "cargo handling", "material handling",
    "water taxi", "ferry", "water ferry", "tug services", "tugboat",
    "barge", "lighterage", "agricultural cleaning", "pratique",

    # WEXMAC adjacent support areas in the attached PWS
    "base operations", "life support", "event support site", "event lot",
    "portable sanitary", "portable shower", "temporary shower", "hand wash station",
    "portable generator", "portable heater", "portable air conditioner",
    "trash removal", "dumpster", "potable water", "non-potable water",
    "laundry services", "billeting", "shelter", "custodial services",
    "pest services", "waste management", "hazardous waste", "medical waste",
    "sewage", "black water", "gray water", "food services", "bottled water",
    "rations",

    # construction equipment / material handling / cranes
    "construction equipment", "material handling equipment", "crane services",
    "mobile crane", "barge lift", "forklift", "k loader", "manlift",
    "scissor lift", "bulldozer", "skid steer", "front end loader",
    "excavator", "fuel truck", "water distributor truck", "dump truck",
    "grader", "asphalt paver", "vibratory roller",

    # lodging, conference, catering, medical logistics, communications, force protection
    "lodging services", "conference services", "catering", "box lunches",
    "berthing barge", "medical logistics", "ambulance services", "veterinary services",
    "communications services", "landline", "cellular phones", "sim cards",
    "wifi internet", "force protection", "security guards", "metal detector",
    "x-ray baggage", "security trailer", "guard shack", "barrier",

    # agency/market signals
    "department of transportation", "dot", "transit authority", "municipal transportation",
    "public transportation", "mass transit", "airport shuttle", "base shuttle",
    "department of veterans affairs", "veterans affairs", "va medical center",
    "department of the interior", "national park service", "nps", "forest service",
    "department of defense", "dod", "installation shuttle", "base transportation",
    "navfac", "us navy", "military sealift command", "msc", "army", "air force",
]

# PSC / classificationCode (signals)
# V-codes generally represent transportation/travel-type services; J/W codes catch
# maintenance/lease/rental-related opportunities that may support vehicle fleet work.
PSCS = [
    # Transportation / travel / relocation / vehicle operations
    "V003", "V112", "V119", "V122", "V129", "V212", "V222",
    "V225", "V226", "V227", "V229", "V999",

    # Logistics, warehousing, cargo, support services
    "R405", "R408", "R499", "R602", "R604", "R605", "R606",
    "R706", "R799", "S216", "S205", "S206", "S208", "S209",

    # Rental/lease and maintenance of vehicles/equipment
    "W023", "W025", "W039", "W099", "J023", "J025", "J039", "J099",

    # Facilities/base support and food/water/life support
    "M1LZ", "S201", "S203", "S211", "S222", "S299", "S203",
]

# NAICS (signals) — Weston Trolley / ground passenger transportation universe
NAICS = [
    # Passenger transportation / Weston Trolley core
    "485113",  # Bus and Other Motor Vehicle Transit Systems
    "485119",  # Other Urban Transit Systems
    "485210",  # Interurban and Rural Bus Transportation
    "485310",  # Taxi and Ridesharing Services
    "485320",  # Limousine Service
    "485410",  # School and Employee Bus Transportation
    "485510",  # Charter Bus Industry
    "485991",  # Special Needs Transportation
    "485999",  # Other Transit and Ground Passenger Transportation
    "487110",  # Scenic and Sightseeing Transportation, Land

    # WEXMAC logistics / cargo / warehousing / support services
    "484110",  # General Freight Trucking, Local
    "484121",  # General Freight Trucking, Long-Distance, Truckload
    "484122",  # General Freight Trucking, Long-Distance, LTL
    "484220",  # Specialized Freight Trucking, Local
    "484230",  # Specialized Freight Trucking, Long-Distance
    "488320",  # Marine Cargo Handling
    "488390",  # Other Support Activities for Water Transportation
    "488410",  # Motor Vehicle Towing
    "488490",  # Other Support Activities for Road Transportation
    "488510",  # Freight Transportation Arrangement
    "488991",  # Packing and Crating
    "488999",  # Other Support Activities for Transportation
    "492110",  # Couriers and Express Delivery Services
    "492210",  # Local Messengers and Local Delivery
    "493110",  # General Warehousing and Storage
    "493190",  # Other Warehousing and Storage
    "541614",  # Process/Physical Distribution/Logistics Consulting Services
    "561210",  # Facilities Support Services
    "561320",  # Temporary Help Services (drivers/operators/labor)
    "561599",  # All Other Travel Arrangement and Reservation Services
    "561920",  # Convention and Trade Show Organizers
    "561990",  # All Other Support Services

    # Rental/lease/maintenance of vehicles and equipment
    "532111",  # Passenger Car Rental
    "532112",  # Passenger Car Leasing
    "532120",  # Truck, Utility Trailer, and RV Rental and Leasing
    "532289",  # Other Consumer Goods Rental
    "532411",  # Commercial Air/Rail/Water Transportation Equipment Rental
    "532412",  # Construction/Mining/Forestry Machinery Rental and Leasing
    "811111",  # General Automotive Repair
    "811118",  # Other Automotive Mechanical/Electrical Repair
    "811198",  # All Other Automotive Repair and Maintenance
    "811310",  # Commercial/Industrial Machinery Repair and Maintenance

    # Base operations / life support / adjacent WEXMAC scope
    "221310",  # Water Supply and Irrigation Systems
    "236220",  # Commercial and Institutional Building Construction
    "238990",  # Other Specialty Trade Contractors
    "541930",  # Translation and Interpretation Services
    "561612",  # Security Guards and Patrol Services
    "561720",  # Janitorial Services
    "561730",  # Landscaping Services
    "561740",  # Carpet and Upholstery Cleaning Services
    "561790",  # Other Services to Buildings and Dwellings
    "562111",  # Solid Waste Collection
    "562112",  # Hazardous Waste Collection
    "562119",  # Other Waste Collection
    "562211",  # Hazardous Waste Treatment and Disposal
    "562991",  # Septic Tank and Related Services
    "562998",  # All Other Miscellaneous Waste Management Services
    "721110",  # Hotels and Motels
    "721214",  # Recreational and Vacation Camps
    "722310",  # Food Service Contractors
]

# Organization codes as LOCAL signals (prefix match on fullParentPathCode)
# Kept broad because Weston Trolley opportunities may come from transportation,
# installation/base support, parks/visitor services, VA, and municipal-adjacent agencies.
ORG_CODES = {
    "DOT": "069",
    "VA": "036",
    "DOI": "014",
    "NPS": "014103",
    "USDA": "012",
    "DOD": "097",
    "GSA": "047",
}

# Place of performance: US states/territories (signals)
# Preserves the existing script regions and adds the previously discussed upper-midwest /
# mountain states, plus the tier 3 home-region states (MD, TN, WV, KY, DC, GA).
POP_STATES = sorted(set([
    "NC", "SC", "VA", "CO", "PA", "PR", "GU",
    "MN", "ND", "SD", "WI",
    "MD", "TN", "WV", "KY", "DC", "GA",
]))

# International PoP targets (signals)
# South Asia, South America, Central America, and Caribbean coverage. Country names
# are matched locally in title/description/agency text after structured searches return.
POP_COUNTRIES = sorted(set([
    # South Asia
    "Afghanistan", "Bangladesh", "Bhutan", "India", "Maldives",
    "Nepal", "Pakistan", "Sri Lanka",

    # South America
    "Argentina", "Bolivia", "Brazil", "Chile", "Colombia", "Ecuador",
    "Guyana", "Paraguay", "Peru", "Suriname", "Uruguay", "Venezuela",

    # Central America
    "Belize", "Costa Rica", "El Salvador", "Guatemala", "Honduras",
    "Nicaragua", "Panama",

    # Caribbean
    "Anguilla", "Antigua", "Antigua and Barbuda", "Aruba", "Bahamas",
    "Barbados", "Bermuda", "Bonaire", "British Virgin Islands",
    "Cayman Islands", "Cuba", "Curacao", "Dominica", "Dominican Republic",
    "Grenada", "Guadeloupe", "Haiti", "Jamaica", "Martinique",
    "Montserrat", "Puerto Rico", "Saint Barthelemy", "Saint Kitts",
    "Saint Kitts and Nevis", "Saint Lucia", "Saint Martin",
    "Saint Vincent", "Saint Vincent and the Grenadines", "Sint Maarten",
    "Trinidad", "Trinidad and Tobago", "Turks and Caicos",
    "U.S. Virgin Islands", "US Virgin Islands", "Virgin Islands",
]))

# Set-aside signals inferred locally from title/description. Weston qualifies for
# Veteran-Owned Small Business and Small Business set-asides.
SETASIDE_KEYWORDS = [
    "VOSB", "Veteran-Owned", "Veteran Owned", "Veteran-Owned Small Business",
    "Veteran Owned Small Business", "SDVOSB", "Service-Disabled",
    "Service Disabled Veteran Owned", "Service-Disabled Veteran-Owned",
    "Small Business", "Total Small Business", "Small Business Set-Aside",
    "SBSA", "set-aside", "set aside",
]

# Logistics-only terms: overhead bump but NOT automatically an international advantage.
LOGISTICS_TERMS = [
    "guam", "puerto rico", "u.s. virgin islands", "us virgin islands",
    "oconus", "overseas", "remote site", "island", "ferry", "port",
    "customs", "bill of lading", "hazmat", "multi-site",
]

# WEXMAC domain-fit terms from the attached PWS. These boost relevance/profitability
# because Weston supports these logistics areas through the WEXMAC contract vehicle.
WEXMAC_FOCUS_TERMS = [
    "cargo truck", "cargo van", "light duty truck", "covered truck", "flatbed",
    "stake truck", "semi truck", "reefer van", "vehicle rental", "with driver",
    "without driver", "bus-26", "bus-40", "bus-50", "airport transfer",
    "personnel logistic movement", "plms", "warehouse", "hazmat warehouse",
    "portable warehouse", "postage", "loading", "unloading", "customs duty",
    "customs clearance", "bill of lading", "logistics support", "water taxi",
    "water ferry", "tug", "barge", "lighterage", "agricultural cleaning",
    "pratique", "air transport",
    "base operations", "life support", "event support site", "event lot",
    "portable sanitary", "portable shower", "hand wash station", "generator",
    "heater", "air conditioner", "trash removal", "dumpster", "potable water",
    "laundry", "custodial", "pest", "waste management", "food services",
    "construction equipment", "material handling", "crane", "forklift", "manlift",
    "scissor lift", "bulldozer", "excavator", "dump truck",
    "lodging", "conference", "catering", "box lunches", "medical logistics",
    "communications", "cellular", "wifi", "force protection", "security guards",
    "metal detector", "x-ray baggage", "guard shack", "security trailer",
]

# Broader WEXMAC/Weston scoring families. These are intentionally not limited to
# trolley/passenger transport. The PWS spans expeditionary logistics, life support,
# material handling, lodging/catering, medical logistics, force protection,
# communications, food/water/supplies, and transportation by road/water/air.
DOMAIN_FAMILIES = {
    "Weston core passenger transport": {
        "weight": 28,
        "terms": [
            "trolley", "streetcar", "shuttle", "bus service", "bus services",
            "charter bus", "motor coach", "motorcoach", "passenger transportation",
            "ground transportation", "surface transportation", "circulator",
            "visitor transportation", "park shuttle", "airport transfer",
            "airport transfers", "paratransit", "microtransit", "driver services",
            "vehicle operator", "bus operator", "fixed route", "transit operations",
        ],
    },
    "WEXMAC logistics and transportation": {
        "weight": 26,
        "terms": [
            "cargo truck", "cargo van", "light duty truck", "covered truck", "flatbed",
            "stake truck", "semi truck", "tractor trailer", "reefer", "refrigerated van",
            "vehicle rental", "vehicle leasing", "with driver", "without driver",
            "personnel logistic movement", "personnel logistics movement", "plms",
            "logistics support", "movement support", "loading", "unloading",
            "customs clearance", "customs duty", "bill of lading", "freight",
            "drayage", "cargo handling", "postage", "courier", "delivery",
        ],
    },
    "Warehousing and supply chain": {
        "weight": 20,
        "terms": [
            "warehouse", "warehousing", "general warehouse", "hazmat warehouse",
            "portable warehouse", "storage services", "supply chain", "inventory",
            "materials management", "packing", "crating", "distribution",
        ],
    },
    "Water transport and port services": {
        "weight": 20,
        "terms": [
            "water taxi", "water ferry", "ferry", "tug", "tugboat", "tow boat",
            "barge", "lighterage", "marine cargo", "port services", "pier",
            "sealift", "military sealift", "msc", "pratique", "agricultural cleaning",
        ],
    },
    "Base operations and life support": {
        "weight": 18,
        "terms": [
            "base operations", "life support", "event support site", "event lot",
            "portable sanitary", "portable shower", "temporary shower", "hand wash station",
            "generator", "portable generator", "heater", "air conditioner", "cooling",
            "trash removal", "dumpster", "potable water", "non-potable water",
            "laundry", "billeting", "shelter", "custodial", "pest", "waste management",
            "hazardous waste", "medical waste", "gray water", "black water", "sewage",
            "food services", "bottled water", "rations",
        ],
    },
    "Equipment and material handling": {
        "weight": 16,
        "terms": [
            "construction equipment", "material handling", "crane", "mobile crane",
            "forklift", "k loader", "manlift", "scissor lift", "bulldozer", "skid steer",
            "front end loader", "excavator", "fuel truck", "water distributor truck",
            "dump truck", "grader", "asphalt paver", "vibratory roller", "equipment rental",
            "equipment services with operator", "equipment services without operator",
        ],
    },
    "Lodging conference and catering": {
        "weight": 14,
        "terms": [
            "lodging", "hotel", "billeting", "conference services", "catering",
            "box lunches", "food service", "berthing barge", "camp services",
        ],
    },
    "Medical logistics and emergency support": {
        "weight": 12,
        "terms": [
            "medical logistics", "medical facility", "role 1", "role 2", "medevac",
            "air ambulance", "ambulance services", "veterinary services", "medical supplies",
        ],
    },
    "Force protection and communications": {
        "weight": 10,
        "terms": [
            "force protection", "security guard", "security guards", "metal detector",
            "x-ray baggage", "explosive detector", "guard shack", "security trailer",
            "barrier", "radio", "landline", "cellular", "sim cards", "wifi internet",
            "communications services", "internet connection",
        ],
    },
}

BUYER_FIT_TERMS = {
    "DOD / military / installation": ["department of defense", "dod", "army", "navy", "air force", "marine corps", "installation", "base", "military", "navfac", "msc", "military sealift"],
    "Parks / visitor movement": ["national park", "national park service", "nps", "forest service", "recreation area", "visitor center", "concession"],
    "Transportation agencies": ["department of transportation", "dot", "transit authority", "public transportation", "mass transit", "airport"],
    "VA / medical campus": ["veterans affairs", "va medical", "medical center"],
    "Disaster / contingency": ["fema", "disaster", "emergency response", "contingency", "humanitarian", "relief", "evacuation"],
}

SETASIDE_BOOST_TERMS = {
    "VOSB/SDVOSB": ["vosb", "veteran-owned", "veteran owned", "sdvosb", "service-disabled", "service disabled veteran"],
    "Small business": ["small business", "total small business", "small business set-aside", "sbsa", "set-aside", "set aside"],
}

# SAM.gov structured set-aside codes (typeOfSetAside). These are authoritative —
# preferred over text scanning. SDVOSB is the priority target for Weston.
SDVOSB_SETASIDE_CODES = {"SDVOSBC", "SDVOSBS"}   # SDVOSB set-aside, SDVOSB sole source
VOSB_SETASIDE_CODES = {"VSA", "VSS"}             # VOSB set-aside, VOSB sole source
SMALLBIZ_SETASIDE_CODES = {"SBA", "SBP"}         # Total Small Business set-aside / partial

# Priority boost applied to the final score when an SDVOSB set-aside is detected,
# so SDVOSB opportunities rank first while everything else still appears.
SDVOSB_PRIORITY_BOOST = 60.0
VOSB_PRIORITY_BOOST = 25.0

# Tiered home-region boost (moderate, NC peak). Applied as a general score boost on
# top of the existing geo scoring, based on place-of-performance / office state.
# Highest matching tier wins (not additive across tiers).
HOME_REGION_TIERS = {
    18.0: {"NC"},                                    # Tier 1 — home base
    12.0: {"SC", "VA"},                              # Tier 2 — adjacent core
    6.0:  {"MD", "TN", "WV", "KY", "PA", "DC", "GA"},  # Tier 3 — broader region
}

NEGATIVE_FIT_TERMS = [
    "software development", "information technology", "it services", "cybersecurity",
    "research and development", "laboratory", "architect-engineer", "a-e services",
    "sf330", "design-bid-build", "vertical construction", "janitorial only",
]


# Email recipients (prefer env vars so you don't edit code).
# Accepts either REPORT_TO or EMAIL_TO (and REPORT_CC/EMAIL_CC) so it works
# regardless of which secret name is configured. Note: a secret that is SET BUT
# EMPTY returns "" from getenv (not the default), so blanks are coerced to the default.
EMAIL_TO = (
    (os.getenv("REPORT_TO") or "").strip()
    or (os.getenv("EMAIL_TO") or "").strip()
    or "jleister@westontrolley.com"
)
EMAIL_CC = (
    (os.getenv("REPORT_CC") or "").strip()
    or (os.getenv("EMAIL_CC") or "").strip()
)
EMAIL_SUBJECT_BASE = "SAM.gov Opportunities – last 72 hours (Feasibility Ranked)"

TOP_MIN, TOP_MAX = 5, 10
SHORTLIST_MIN, SHORTLIST_MAX = 10, 20

# Performance caps
MAX_PER_JOB = 2000        # max records to pull per query job
MAX_TOTAL_DEDUPED = 6000  # stop early once enough deduped candidates exist
SLEEP_SECONDS = 0.12

# Feasibility scoring weights
# Score = (Profitability / (Complexity + Overhead)) * 10  +  0.15 * RelevanceSignals
FEASIBILITY_MULT = 10.0
RELEVANCE_WEIGHT = 0.15
MAX_RELEVANCE = 10.0


# -----------------------------
# OPTIONAL SMTP CONFIG (send only if SEND_EMAIL=1)
# -----------------------------
SEND_EMAIL = os.getenv("SEND_EMAIL", "0").strip() == "1"
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
FROM_EMAIL = os.getenv("FROM_EMAIL", SMTP_USER)


# -----------------------------
# DATA MODEL
# -----------------------------
@dataclass
class Opportunity:
    noticeId: str
    title: str
    uiLink: str
    postedDate: Optional[str] = None
    responseDeadLine: Optional[str] = None
    type: Optional[str] = None
    baseType: Optional[str] = None
    fullParentPathName: Optional[str] = None
    fullParentPathCode: Optional[str] = None
    naicsCodes: List[str] = field(default_factory=list)
    classificationCode: Optional[str] = None
    active: Optional[str] = None
    office_state: Optional[str] = None
    description_url: Optional[str] = None
    resourceLinks: List[str] = field(default_factory=list)
    typeOfSetAside: Optional[str] = None          # structured set-aside code (e.g. SDVOSBC)
    typeOfSetAsideDescription: Optional[str] = None

    # derived
    why_matched: List[str] = field(default_factory=list)
    description_text: str = ""
    keyword_hits: List[str] = field(default_factory=list)

    ratings: Dict[str, Any] = field(default_factory=dict)
    evidence: List[str] = field(default_factory=list)
    next_step: str = ""
    score: float = 0.0
    feasibility: float = 0.0


# -----------------------------
# HELPERS
# -----------------------------
def require_env(name: str) -> str:
    v = os.getenv(name)
    if not v:
        raise RuntimeError(f"Missing required env var: {name}")
    return v


def mmddyyyy(d: dt.date) -> str:
    return d.strftime("%m/%d/%Y")


def parse_iso_date(iso_dt: str) -> Optional[dt.datetime]:
    try:
        return dt.datetime.fromisoformat(iso_dt.replace("Z", "+00:00"))
    except Exception:
        return None


def sam_search(api_key: str, params: Dict[str, Any], timeout: int = 60) -> Dict[str, Any]:
    q = dict(params)
    q["api_key"] = api_key
    r = requests.get(SAM_SEARCH_URL, params=q, timeout=timeout)
    if r.status_code != 200:
        raise RuntimeError(f"SAM API error {r.status_code}: {r.text[:800]}")
    return r.json()


def sam_fetch_description(desc_url: str, timeout: int = 60) -> str:
    """
    desc_url often is a v1 noticedesc endpoint; public but can be flaky.
    Best-effort pull.
    """
    if not desc_url:
        return ""
    try:
        r = requests.get(desc_url, timeout=timeout)
        if r.status_code != 200:
            return ""
        ctype = r.headers.get("content-type", "")
        if "application/json" in ctype:
            data = r.json()
            for k in ("description", "noticeDesc", "data"):
                if isinstance(data.get(k), str):
                    return data[k][:20000]
            return json.dumps(data)[:20000]
        return r.text[:20000]
    except Exception:
        return ""


def normalize(item: Dict[str, Any]) -> Opportunity:
    office = item.get("officeAddress") or {}
    return Opportunity(
        noticeId=item.get("noticeId") or "",
        title=item.get("title") or "",
        uiLink=item.get("uiLink") or "",
        postedDate=item.get("postedDate"),
        responseDeadLine=item.get("responseDeadLine") or item.get("responseDeadline"),
        type=item.get("type"),
        baseType=item.get("baseType"),
        fullParentPathName=item.get("fullParentPathName"),
        fullParentPathCode=item.get("fullParentPathCode"),
        naicsCodes=item.get("naicsCodes") or ([item["naicsCode"]] if item.get("naicsCode") else []),
        classificationCode=item.get("classificationCode"),
        active=item.get("active"),
        office_state=office.get("state"),
        description_url=item.get("description"),
        resourceLinks=item.get("resourceLinks") or [],
        typeOfSetAside=item.get("typeOfSetAside"),
        typeOfSetAsideDescription=item.get("typeOfSetAsideDescription"),
    )


def hard_filters_ok(opp: Opportunity, today: dt.date, due_max: dt.date) -> bool:
    if ACTIVE_ONLY and (opp.active or "").lower() != "yes":
        return False

    # due within next year (best-effort)
    if opp.responseDeadLine:
        due_dt = parse_iso_date(opp.responseDeadLine)
        if due_dt and not (today <= due_dt.date() <= due_max):
            return False

    return True


def keyword_hits_local(text: str) -> List[str]:
    t = (text or "").lower()
    hits = []
    for kw in KEYWORDS:
        k = kw.lower()
        if k and k in t:
            hits.append(kw)
    # de-dupe but keep order
    out: List[str] = []
    seen = set()
    for h in hits:
        if h not in seen:
            out.append(h)
            seen.add(h)
    return out


def add_job_tag(opp: Opportunity, job_tag: str) -> None:
    tag = f"signal:{job_tag}"
    if tag not in opp.why_matched:
        opp.why_matched.append(tag)


def add_structural_reasons(opp: Opportunity) -> None:
    # PSC
    if opp.classificationCode and opp.classificationCode.upper() in {p.upper() for p in PSCS}:
        opp.why_matched.append(f"PSC:{opp.classificationCode}")

    # NAICS
    naics_hit = sorted(set(opp.naicsCodes).intersection(NAICS))
    if naics_hit:
        opp.why_matched.append("NAICS:" + ",".join(naics_hit))

    # Org codes (prefix match on fullParentPathCode)
    if opp.fullParentPathCode:
        for name, code in ORG_CODES.items():
            if str(opp.fullParentPathCode).startswith(str(code)):
                opp.why_matched.append(f"Org:{name}({code})")
                break

    # Office state weak signal
    if opp.office_state and opp.office_state.upper() in set(POP_STATES):
        opp.why_matched.append(f"OfficeState:{opp.office_state.upper()}")


# -----------------------------
# RATING LOGIC
# -----------------------------
def _hits(text: str, terms: List[str]) -> List[str]:
    out: List[str] = []
    for term in terms:
        if term.lower() in text:
            out.append(term)
    return list(dict.fromkeys(out))


def _family_scores(text: str) -> Tuple[Dict[str, int], Dict[str, List[str]]]:
    """Return weighted domain family scores and matching evidence terms."""
    scores: Dict[str, int] = {}
    evidence: Dict[str, List[str]] = {}
    for family, cfg in DOMAIN_FAMILIES.items():
        terms = cfg["terms"]
        matches = _hits(text, terms)
        if matches:
            # Do not let one repeated family dominate just because many near-synonyms appear.
            scores[family] = int(cfg["weight"]) + min(12, 3 * (len(matches) - 1))
            evidence[family] = matches[:8]
    return scores, evidence


def _buyer_scores(text: str) -> Tuple[int, List[str]]:
    score = 0
    evidence: List[str] = []
    for family, terms in BUYER_FIT_TERMS.items():
        matches = _hits(text, terms)
        if matches:
            score += 6
            evidence.append(f"{family}: {', '.join(matches[:3])}")
    return min(score, 18), evidence[:4]


def home_region_boost(opp: Opportunity) -> Tuple[float, Optional[str]]:
    """
    Return the highest matching home-region tier boost and the state that triggered it.
    Checks the office state first, then place-of-performance state text signals.
    Highest tier wins (boosts are not stacked across tiers).
    """
    candidates = set()
    if opp.office_state:
        candidates.add(opp.office_state.strip().upper())

    # Place-of-performance / description text signals (e.g. "NC", "North Carolina" abbrev).
    blob = " ".join([
        opp.title or "",
        opp.description_text or "",
        " ".join(opp.why_matched or []),
    ]).upper()

    best_boost = 0.0
    best_state: Optional[str] = None
    for boost, states in HOME_REGION_TIERS.items():
        for st in states:
            # office state match, or whitespace-delimited state token in text
            if st in candidates or f" {st} " in f" {blob} " or f" {st}," in blob:
                if boost > best_boost:
                    best_boost = boost
                    best_state = st
    return best_boost, best_state


def classify_setaside(opp: Opportunity) -> Optional[str]:
    """
    Return 'SDVOSB', 'VOSB', or 'SmallBiz' based on SAM's structured set-aside code,
    falling back to the description text only when the structured code is absent.
    """
    code = (opp.typeOfSetAside or "").strip().upper()
    if code in SDVOSB_SETASIDE_CODES:
        return "SDVOSB"
    if code in VOSB_SETASIDE_CODES:
        return "VOSB"
    if code in SMALLBIZ_SETASIDE_CODES:
        return "SmallBiz"

    # Fallback: description text. Check SDVOSB first since it is the priority.
    blob = " ".join([
        opp.typeOfSetAsideDescription or "",
        opp.title or "",
        opp.description_text or "",
    ]).lower()
    if any(t in blob for t in ["sdvosb", "service-disabled veteran", "service disabled veteran"]):
        return "SDVOSB"
    if any(t in blob for t in ["vosb", "veteran-owned", "veteran owned"]):
        return "VOSB"
    if any(t in blob for t in ["small business set-aside", "total small business", "8(a)", "hubzone", "wosb", "edwosb"]):
        return "SmallBiz"
    return None


def _setaside_score(text: str) -> Tuple[int, List[str]]:
    score = 0
    evidence: List[str] = []
    for label, terms in SETASIDE_BOOST_TERMS.items():
        matches = _hits(text, terms)
        if matches:
            score += 10 if "VOSB" in label else 7
            evidence.append(f"{label} signal: {', '.join(matches[:3])}")
    return min(score, 17), evidence


def estimate_ratings(opp: Opportunity) -> None:
    """
    Broader Weston/WEXMAC fit model.

    Instead of only rewarding trolley/shuttle words, this scores across the full WEXMAC
    PWS scope: transportation, warehousing, base/life support, equipment/material
    handling, lodging/catering, medical logistics, communications, force protection,
    food/water/supplies, and international/contingency logistics.
    """
    text = f"{opp.title}\n{opp.type}\n{opp.baseType}\n{opp.fullParentPathName}\n{opp.description_text}".lower()

    family_scores, family_evidence = _family_scores(text)
    buyer_score, buyer_evidence = _buyer_scores(text)
    setaside_score, setaside_evidence = _setaside_score(text)

    country_hits = _hits(text, [c.lower() for c in POP_COUNTRIES])
    state_hits = _hits(text, [s.lower() for s in POP_STATES])
    remote_hits = _hits(text, LOGISTICS_TERMS)
    negative_hits = _hits(text, NEGATIVE_FIT_TERMS)

    domain_fit = min(70, sum(family_scores.values()))
    geo_fit = 0
    geo_evidence: List[str] = []
    if country_hits:
        geo_fit += 12
        geo_evidence.append("Target country/region signal: " + ", ".join(country_hits[:5]))
    if opp.office_state and opp.office_state.upper() in set(POP_STATES):
        geo_fit += 6
        geo_evidence.append(f"Target office/place state signal: {opp.office_state.upper()}")
    elif state_hits:
        geo_fit += 4
        geo_evidence.append("Target state text signal: " + ", ".join(state_hits[:5]))
    if remote_hits:
        geo_fit += 4
        geo_evidence.append("Remote/OCONUS/logistics complexity signal: " + ", ".join(remote_hits[:4]))
    geo_fit = min(18, geo_fit)

    # Convert broad opportunity fit into ratings used by the legacy email columns.
    total_fit = domain_fit + buyer_score + setaside_score + geo_fit - min(18, 5 * len(negative_hits))

    complexity = 3
    overhead = 3
    profitability = 3

    # Higher domain fit and set-aside fit improves profit potential.
    if domain_fit >= 45:
        profitability += 2
    elif domain_fit >= 24:
        profitability += 1
    if setaside_score:
        profitability += 1
    if buyer_score >= 12:
        profitability += 1

    # Some WEXMAC families are inherently heavier to execute.
    heavy_families = {
        "Base operations and life support",
        "Equipment and material handling",
        "Medical logistics and emergency support",
        "Force protection and communications",
        "Water transport and port services",
    }
    if any(f in family_scores for f in heavy_families):
        complexity += 1
        overhead += 1
    if country_hits or remote_hits:
        overhead += 1
    if any(w in text for w in ["nationwide", "multi-site", "multiple locations", "24/7", "twenty four seven", "classified", "secret", "top secret"]):
        complexity += 1
        overhead += 1
    if negative_hits:
        complexity += 1

    # Keep routine passenger/vehicle rental work from being unfairly penalized.
    if any(f in family_scores for f in ["Weston core passenger transport", "WEXMAC logistics and transportation", "Warehousing and supply chain"]):
        complexity -= 1

    complexity = max(1, min(5, complexity))
    overhead = max(1, min(5, overhead))
    profitability = max(1, min(5, profitability))

    evidence: List[str] = []
    if family_scores:
        top_families = sorted(family_scores.items(), key=lambda x: x[1], reverse=True)[:4]
        for fam, pts in top_families:
            evidence.append(f"{fam} fit (+{pts}): {', '.join(family_evidence[fam][:5])}")
    evidence.extend(setaside_evidence)
    evidence.extend(buyer_evidence)
    evidence.extend(geo_evidence)
    if negative_hits:
        evidence.append("Possible off-scope/low-fit signal: " + ", ".join(negative_hits[:4]))

    opp.ratings = {
        "complexity": complexity,
        "profitability": profitability,
        "overhead": overhead,
        "domain_fit": domain_fit,
        "buyer_fit": buyer_score,
        "setaside_fit": setaside_score,
        "setaside_class": classify_setaside(opp),
        "geo_fit": geo_fit,
        "total_fit": max(0, total_fit),
    }
    opp.evidence = evidence[:7]
    opp.next_step = (
        "Review the notice/attachments against Weston prime/sub role. Confirm WEXMAC family fit, set-aside status, "
        "place of performance, vehicle/equipment/labor requirements, mobilization burden, insurance/licensing, and bid deadline."
    )


def compute_score(opp: Opportunity) -> float:
    """
    Broad business-fit ranking.

    New score is not just feasibility. It gives primary weight to Weston/WEXMAC domain fit,
    then set-aside eligibility, buyer fit, geography, and feasibility. This avoids burying
    broad logistics/life-support opportunities just because they are not trolley/shuttle jobs.
    """
    c = float(opp.ratings.get("complexity", 3))
    o = float(opp.ratings.get("overhead", 3))
    p = float(opp.ratings.get("profitability", 3))
    feasibility = p / max(1.0, c + o)
    opp.feasibility = feasibility

    domain_fit = float(opp.ratings.get("domain_fit", 0))
    buyer_fit = float(opp.ratings.get("buyer_fit", 0))
    setaside_fit = float(opp.ratings.get("setaside_fit", 0))
    geo_fit = float(opp.ratings.get("geo_fit", 0))
    rel = min(MAX_RELEVANCE, float(len(set(opp.why_matched))))

    # SDVOSB priority boost: surface SDVOSB set-asides first, without excluding others.
    setaside_class = opp.ratings.get("setaside_class")
    priority_boost = 0.0
    if setaside_class == "SDVOSB":
        priority_boost = SDVOSB_PRIORITY_BOOST
    elif setaside_class == "VOSB":
        priority_boost = VOSB_PRIORITY_BOOST

    # Tiered home-region boost (stacks on top of existing geo_fit), NC peak.
    region_boost, region_state = home_region_boost(opp)
    if region_boost:
        opp.ratings["home_region_boost"] = region_boost
        opp.ratings["home_region_state"] = region_state

    # 0-100ish scale. Domain fit dominates; feasibility still matters but is no longer the gatekeeper.
    return domain_fit + buyer_fit + setaside_fit + geo_fit + (feasibility * 12.0) + (rel * 0.75) + priority_boost + region_boost


def get_setaside_label(opp: Opportunity) -> str:
    """Return a compact set-aside label, preferring SAM's structured set-aside code."""
    # Structured classification (authoritative) takes priority.
    cls = opp.ratings.get("setaside_class") if opp.ratings else None
    if cls == "SDVOSB":
        return "SDVOSB"
    if cls == "VOSB":
        return "VOSB"

    # Fallback to local text signals.
    text = " ".join([
        opp.title or "",
        opp.typeOfSetAsideDescription or "",
        opp.description_text or "",
        " ".join(opp.why_matched or []),
    ]).lower()
    labels = []
    if any(t in text for t in ["sdvosb", "service-disabled", "service disabled veteran"]):
        labels.append("SDVOSB")
    elif any(t in text for t in ["vosb", "veteran-owned", "veteran owned"]):
        labels.append("VOSB")
    if any(t in text for t in ["small business", "total small business", "small business set-aside", "sbsa", "set-aside", "set aside"]):
        labels.append("Small Business")
    return ", ".join(labels) if labels else "Not identified"


def get_location_label(opp: Opportunity) -> str:
    """Return a compact location label from office_state plus country/geography signals."""
    pieces = []
    if opp.office_state:
        pieces.append(opp.office_state)
    why = " | ".join(opp.why_matched or [])
    country_hits = []
    for country in POP_COUNTRIES:
        if country.lower() in why.lower():
            country_hits.append(country)
    if country_hits:
        pieces.extend(country_hits[:3])
    return ", ".join(dict.fromkeys(pieces)) if pieces else "Not specified"


def get_match_summary(opp: Opportunity, max_items: int = 4) -> str:
    """Create a short human-readable reason string for the email report."""
    items = []
    for ev in opp.evidence or []:
        if ev and ev not in items:
            items.append(ev)
    for why in opp.why_matched or []:
        if why and why not in items:
            items.append(why)
    if not items and opp.keyword_hits:
        items = ["keywords: " + ", ".join(opp.keyword_hits[:max_items])]
    return "; ".join(items[:max_items]) if items else "Matched Weston/WEXMAC search signals"


def build_email(top: List[Opportunity], shortlist: List[Opportunity], as_of: dt.datetime) -> str:
    """Plain-text version of the daily BD report. The HTML version is used for email clients."""
    all_reported = top + shortlist
    vosb_count = sum(1 for o in all_reported if "VOSB" in get_setaside_label(o) or "SDVOSB" in get_setaside_label(o))
    sb_count = sum(1 for o in all_reported if "Small Business" in get_setaside_label(o))

    lines: List[str] = []
    lines.append(f"To: {EMAIL_TO}")
    if EMAIL_CC.strip():
        lines.append(f"Cc: {EMAIL_CC}")
    lines.append(f"Subject: Weston Trolley / WEXMAC Daily Opportunities — {as_of:%b %d, %Y}")
    lines.append("")
    lines.append("Good afternoon,")
    lines.append("")
    lines.append(f"Here is today's SAM.gov opportunity scan for Weston Trolley and WEXMAC logistics focus areas.")
    lines.append(f"Search window: last ~{POSTED_WINDOW_HOURS} hours")
    lines.append(f"Top opportunities: {len(top)}")
    lines.append(f"Next-best shortlist: {len(shortlist)}")
    lines.append(f"VOSB/SDVOSB signals in report: {vosb_count}")
    lines.append(f"Small Business signals in report: {sb_count}")
    lines.append("")

    lines.append("HIGH PRIORITY / TOP OPPORTUNITIES")
    lines.append("=" * 72)
    if not top:
        lines.append("No top opportunities found for this run.")
    for i, opp in enumerate(top, 1):
        lines.append(f"{i}) {opp.title}")
        lines.append(f"   - Score: {opp.score:.1f} | Feasibility: {opp.feasibility:.2f}")
        lines.append(f"   - Agency/Office: {opp.fullParentPathName or '—'}")
        lines.append(f"   - Location: {get_location_label(opp)}")
        lines.append(f"   - Set-aside: {get_setaside_label(opp)}")
        lines.append(f"   - Posted: {opp.postedDate or '—'} | Due: {opp.responseDeadLine or '—'}")
        lines.append(f"   - NAICS: {', '.join(opp.naicsCodes) if opp.naicsCodes else '—'} | PSC: {opp.classificationCode or '—'}")
        lines.append(f"   - Why it matters: {get_match_summary(opp)}")
        lines.append(f"   - Next step: {opp.next_step}")
        lines.append(f"   - SAM link: {opp.uiLink}")
        lines.append("")

    lines.append("OTHER STRONG MATCHES")
    lines.append("=" * 72)
    if not shortlist:
        lines.append("No shortlist opportunities found for this run.")
    for opp in shortlist:
        lines.append(
            f"- {opp.title} | Score {opp.score:.1f} | {get_location_label(opp)} | "
            f"{get_setaside_label(opp)} | Due {opp.responseDeadLine or '—'}"
        )
        lines.append(f"  {opp.uiLink}")

    lines.append("")
    lines.append("Full ranked results are attached in Excel/CSV.")
    return "\n".join(lines)


def opp_to_row(opp: Opportunity, rank_group: str = "") -> Dict[str, Any]:
    """Flatten an opportunity for CSV/XLSX output."""
    return {
        "rank_group": rank_group,
        "score": round(float(opp.score or 0), 3),
        "feasibility": round(float(opp.feasibility or 0), 3),
        "complexity": opp.ratings.get("complexity", ""),
        "profitability": opp.ratings.get("profitability", ""),
        "overhead": opp.ratings.get("overhead", ""),
        "title": opp.title,
        "notice_type": opp.type,
        "posted_date": opp.postedDate,
        "response_deadline": opp.responseDeadLine,
        "agency_office": opp.fullParentPathName,
        "naics": ", ".join(opp.naicsCodes or []),
        "psc": opp.classificationCode or "",
        "office_state": opp.office_state or "",
        "keyword_hits": ", ".join(opp.keyword_hits[:12]),
        "why_matched": "; ".join(dict.fromkeys(opp.why_matched)),
        "evidence": " | ".join(opp.evidence),
        "next_step": opp.next_step,
        "notice_id": opp.noticeId,
        "sam_link": opp.uiLink,
        "attachment_count": len(opp.resourceLinks or []),
    }


def write_results_csv(scored: List[Opportunity], top_ids: set, shortlist_ids: set, as_of: dt.datetime) -> str:
    filename = f"sam_results_weston_wexmac_{as_of:%Y-%m-%d}.csv"
    fieldnames = list(opp_to_row(scored[0] if scored else Opportunity('', '', '')).keys())
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for opp in scored:
            group = "Top" if opp.noticeId in top_ids else ("Shortlist" if opp.noticeId in shortlist_ids else "All Matches")
            writer.writerow(opp_to_row(opp, group))
    return filename


def write_results_xlsx(scored: List[Opportunity], top_ids: set, shortlist_ids: set, as_of: dt.datetime) -> Optional[str]:
    """Write an Excel workbook if openpyxl is installed; otherwise skip gracefully."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    filename = f"sam_results_weston_wexmac_{as_of:%Y-%m-%d}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "SAM Results"

    rows = []
    for opp in scored:
        group = "Top" if opp.noticeId in top_ids else ("Shortlist" if opp.noticeId in shortlist_ids else "All Matches")
        rows.append(opp_to_row(opp, group))

    headers = list(rows[0].keys()) if rows else list(opp_to_row(Opportunity('', '', '')).keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 14, "B": 10, "C": 12, "D": 10, "E": 12, "F": 10,
        "G": 55, "H": 14, "I": 14, "J": 22, "K": 45, "L": 20,
        "M": 12, "N": 14, "O": 35, "P": 50, "Q": 50, "R": 55,
        "S": 18, "T": 55, "U": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Generated", as_of.strftime("%Y-%m-%d %H:%M")])
    summary.append(["Total scored matches", len(scored)])
    summary.append(["Top opportunities", len(top_ids)])
    summary.append(["Shortlist opportunities", len(shortlist_ids)])
    summary.append(["Search window hours", POSTED_WINDOW_HOURS])
    for cell in summary[1]:
        cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 28

    wb.save(filename)
    return filename


def build_html_email(top: List[Opportunity], shortlist: List[Opportunity], as_of: dt.datetime) -> str:
    """Build a clean HTML BD report for the daily email."""
    def esc(x: Any) -> str:
        return html.escape(str(x or ""))

    all_reported = top + shortlist
    vosb_count = sum(1 for o in all_reported if "VOSB" in get_setaside_label(o) or "SDVOSB" in get_setaside_label(o))
    sb_count = sum(1 for o in all_reported if "Small Business" in get_setaside_label(o))

    def opportunity_card(i: int, opp: Opportunity) -> str:
        return f"""
        <tr>
          <td style="vertical-align:top;padding:8px;border-bottom:1px solid #ddd;">{i}</td>
          <td style="vertical-align:top;padding:8px;border-bottom:1px solid #ddd;">
            <div style="font-weight:700;font-size:14px;">{esc(opp.title)}</div>
            <div style="margin-top:4px;"><a href="{esc(opp.uiLink)}">Open in SAM.gov</a></div>
            <div style="margin-top:6px;color:#444;"><strong>Why it matters:</strong> {esc(get_match_summary(opp))}</div>
          </td>
          <td style="vertical-align:top;padding:8px;border-bottom:1px solid #ddd;">{esc(opp.fullParentPathName or '—')}</td>
          <td style="vertical-align:top;padding:8px;border-bottom:1px solid #ddd;">{esc(get_location_label(opp))}</td>
          <td style="vertical-align:top;padding:8px;border-bottom:1px solid #ddd;">{esc(get_setaside_label(opp))}</td>
          <td style="vertical-align:top;padding:8px;border-bottom:1px solid #ddd;">{esc(opp.responseDeadLine or '—')}</td>
          <td style="vertical-align:top;padding:8px;border-bottom:1px solid #ddd;text-align:right;">{opp.score:.1f}<br><span style="color:#666;font-size:12px;">Feas {opp.feasibility:.2f}</span></td>
        </tr>
        """

    top_rows = "".join(opportunity_card(i, opp) for i, opp in enumerate(top, 1))
    if not top_rows:
        top_rows = "<tr><td colspan='7' style='padding:10px;'>No high-priority opportunities found for this run.</td></tr>"

    shortlist_items = "".join(
        f"""
        <li style="margin-bottom:8px;">
          <strong>{esc(opp.title)}</strong><br>
          Score {opp.score:.1f} | {esc(get_location_label(opp))} | {esc(get_setaside_label(opp))} | Due {esc(opp.responseDeadLine or '—')}<br>
          <a href="{esc(opp.uiLink)}">Open in SAM.gov</a>
        </li>
        """
        for opp in shortlist[:15]
    ) or "<li>No shortlist opportunities found for this run.</li>"

    return f"""
    <html>
    <body style="font-family:Arial, Helvetica, sans-serif;color:#222;line-height:1.35;">
      <h2 style="margin-bottom:4px;">Weston Trolley / WEXMAC Daily Opportunities</h2>
      <p style="margin-top:0;color:#555;">Generated {as_of:%b %d, %Y %H:%M}. Search window: last ~{POSTED_WINDOW_HOURS} hours.</p>

      <table cellspacing="0" cellpadding="0" style="border-collapse:collapse;margin:12px 0 18px 0;">
        <tr>
          <td style="padding:8px 18px 8px 0;"><strong>Total in email</strong><br>{len(all_reported)}</td>
          <td style="padding:8px 18px 8px 0;"><strong>High priority</strong><br>{len(top)}</td>
          <td style="padding:8px 18px 8px 0;"><strong>VOSB/SDVOSB signals</strong><br>{vosb_count}</td>
          <td style="padding:8px 18px 8px 0;"><strong>Small Business signals</strong><br>{sb_count}</td>
        </tr>
      </table>

      <h3 style="margin-bottom:8px;">High Priority / Top Opportunities</h3>
      <table cellspacing="0" cellpadding="0" style="border-collapse:collapse;width:100%;font-size:13px;">
        <thead>
          <tr style="background:#f2f2f2;">
            <th style="text-align:left;padding:8px;border-bottom:2px solid #ccc;">#</th>
            <th style="text-align:left;padding:8px;border-bottom:2px solid #ccc;">Opportunity</th>
            <th style="text-align:left;padding:8px;border-bottom:2px solid #ccc;">Agency / Office</th>
            <th style="text-align:left;padding:8px;border-bottom:2px solid #ccc;">Location</th>
            <th style="text-align:left;padding:8px;border-bottom:2px solid #ccc;">Set-aside</th>
            <th style="text-align:left;padding:8px;border-bottom:2px solid #ccc;">Due</th>
            <th style="text-align:right;padding:8px;border-bottom:2px solid #ccc;">Score</th>
          </tr>
        </thead>
        <tbody>{top_rows}</tbody>
      </table>

      <h3 style="margin-top:22px;">Other Strong Matches</h3>
      <ol>{shortlist_items}</ol>

      <p style="margin-top:18px;">The attached Excel/CSV files include the full ranked result set with scores, NAICS/PSC, match signals, next steps, and SAM.gov links.</p>
    </body>
    </html>
    """


def _parse_addrs(raw: str) -> List[str]:
    """Split a recipient string on commas/semicolons, strip, and drop blanks."""
    if not raw:
        return []
    parts = re.split(r"[;,]", raw)
    return [p.strip() for p in parts if p.strip()]


def send_email(subject: str, body: str, html_body: Optional[str] = None, attachments: Optional[List[str]] = None) -> None:
    if not SEND_EMAIL:
        return
    if not SMTP_USER or not SMTP_PASS:
        raise RuntimeError("SEND_EMAIL=1 but SMTP_USER/SMTP_PASS not set.")

    to_addrs = _parse_addrs(EMAIL_TO)
    cc_addrs = _parse_addrs(EMAIL_CC)
    all_recipients = to_addrs + cc_addrs
    if not all_recipients:
        raise RuntimeError(
            "No valid email recipients. Set the REPORT_TO secret to one or more "
            "comma-separated addresses (e.g. 'a@x.com, b@y.com'). REPORT_CC is optional."
        )

    msg = EmailMessage()
    msg["From"] = FROM_EMAIL
    msg["To"] = ", ".join(to_addrs)
    if cc_addrs:
        msg["Cc"] = ", ".join(cc_addrs)
    msg["Subject"] = subject
    msg.set_content(body)
    if html_body:
        msg.add_alternative(html_body, subtype="html")

    for path_str in attachments or []:
        path = Path(path_str)
        if not path.exists():
            continue
        ctype, _ = mimetypes.guess_type(str(path))
        maintype, subtype = (ctype or "application/octet-stream").split("/", 1)
        msg.add_attachment(path.read_bytes(), maintype=maintype, subtype=subtype, filename=path.name)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60) as s:
        s.ehlo()
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.send_message(msg, from_addr=FROM_EMAIL, to_addrs=all_recipients)


# -----------------------------
# MAIN
# -----------------------------
def run() -> int:
    api_key = require_env("SAM_API_KEY")

    now = dt.datetime.now()
    today = now.date()
    posted_from = (now - dt.timedelta(hours=POSTED_WINDOW_HOURS)).date()
    posted_to = today
    due_max = today + dt.timedelta(days=365)

    base = {
        "postedFrom": mmddyyyy(posted_from),
        "postedTo": mmddyyyy(posted_to),
        "ptype": ",".join(NOTICE_TYPES),
        "active": "Yes" if ACTIVE_ONLY else "No",
        "limit": 1000,
    }

    # Jobs: structured filters only (no q=)
    jobs: List[Tuple[str, Dict[str, Any]]] = []

    # PoP states
    for st in POP_STATES:
        jobs.append((f"state:{st}", {**base, "state": st}))

    # NAICS: single-code calls
    for code in NAICS:
        jobs.append((f"naics:{code}", {**base, "ncode": code}))

    # PSC: single-code calls
    for code in PSCS:
        jobs.append((f"psc:{code}", {**base, "ccode": code}))

    seen: Dict[str, Opportunity] = {}
    total_calls = 0
    job_counts: Dict[str, int] = {}

    for job_name, params in jobs:
        offset = 0
        while True:
            p = dict(params)
            p["offset"] = offset
            data = sam_search(api_key, p)
            total_calls += 1

            items = data.get("opportunitiesData") or []
            job_counts[job_name] = job_counts.get(job_name, 0) + len(items)

            if not items:
                break

            for item in items:
                opp = normalize(item)
                if not opp.noticeId:
                    continue
                if opp.noticeId not in seen:
                    seen[opp.noticeId] = opp
                add_job_tag(seen[opp.noticeId], job_name)

            if len(items) < int(p["limit"]):
                break

            offset += int(p["limit"])
            if offset > 10000:
                break
            if offset >= MAX_PER_JOB:
                break
            if len(seen) >= MAX_TOTAL_DEDUPED:
                break

            time.sleep(SLEEP_SECONDS)

        if len(seen) >= MAX_TOTAL_DEDUPED:
            break

    scored: List[Opportunity] = []
    for opp in seen.values():
        if not hard_filters_ok(opp, today=today, due_max=due_max):
            continue

        add_structural_reasons(opp)

        # Description (best-effort)
        opp.description_text = sam_fetch_description(opp.description_url)
        text_lower = (opp.description_text or "").lower()

        # Local keyword hits
        opp.keyword_hits = keyword_hits_local(f"{opp.title} {opp.fullParentPathName or ''} {opp.description_text}")
        if opp.keyword_hits:
            opp.why_matched.append("keyword:text " + ", ".join(opp.keyword_hits[:4]))

        # International country mentions (signals)
        for ctry in POP_COUNTRIES:
            if ctry.lower() in text_lower:
                opp.why_matched.append(f"POP:{ctry}")

        # WEXMAC focus mentions (signals)
        wexmac_hits = [w for w in WEXMAC_FOCUS_TERMS if w.lower() in text_lower]
        if wexmac_hits:
            opp.why_matched.append("WEXMAC:" + ", ".join(wexmac_hits[:4]))

        # Set-aside mentions (signals)
        if any(s.lower() in text_lower for s in [x.lower() for x in SETASIDE_KEYWORDS]):
            opp.why_matched.append("set-aside:text veteran/small business")

        estimate_ratings(opp)
        opp.score = compute_score(opp)
        scored.append(opp)

    scored.sort(key=lambda x: x.score, reverse=True)

    top = scored[:TOP_MAX]
    if len(top) < TOP_MIN:
        top = scored[:max(TOP_MIN, len(scored))]

    top_ids = {o.noticeId for o in top}
    remaining = [o for o in scored if o.noticeId not in top_ids]

    shortlist = remaining[:SHORTLIST_MAX]
    if len(shortlist) < SHORTLIST_MIN:
        shortlist = remaining[:max(SHORTLIST_MIN, len(remaining))]

    top_ids = {o.noticeId for o in top}
    shortlist_ids = {o.noticeId for o in shortlist}

    csv_path = write_results_csv(scored, top_ids, shortlist_ids, now)
    xlsx_path = write_results_xlsx(scored, top_ids, shortlist_ids, now)

    email_text = build_email(top, shortlist, now)
    email_html = build_html_email(top, shortlist, now)

    with open("email_draft.txt", "w", encoding="utf-8") as f:
        f.write(email_text)
    with open("email_draft.html", "w", encoding="utf-8") as f:
        f.write(email_html)

    print(email_text)

    subject = f"Weston Trolley / WEXMAC Daily Opportunities ({len(scored)} matches | {len(top)} high priority) — {now:%b %d, %Y}"
    attachments = [p for p in [xlsx_path, csv_path] if p]
    send_email(subject, email_text, html_body=email_html, attachments=attachments)

    print(f"[INFO] Wrote spreadsheet files: {', '.join(attachments)}", file=sys.stderr)

    print(
        f"\n[INFO] API calls: {total_calls} | Deduped candidates: {len(seen)} | Scored candidates: {len(scored)} | Top: {len(top)} | Shortlist: {len(shortlist)} | SEND_EMAIL={int(SEND_EMAIL)}",
        file=sys.stderr
    )
    for k in sorted(job_counts, key=lambda x: (-job_counts[x], x)):
        print(f"[JOB] {k}: {job_counts[k]}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(run())
